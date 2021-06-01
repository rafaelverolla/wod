from tkinter import Button, Toplevel, Entry, Label, Tk, messagebox, Menu, Frame, END
from tkinter import ttk
import copy
import os
import json
import openpyxl

#Objective: make a WoD sheet manager that let you create an initial char, save it permanently, export/import to/from an excel file.
#The ui must be intuitive. Also more then 1 type of character must be accepted (Mage, Werewolf and Vampire, at least).
#The program must do all the  math automatically (Points distribution, freebies and experience), and let the user know if something is wrong.
#A ST mode should be done, where the user can do a sheet without mathematical limits, and the NPC sheet exported should be simple, with only what the npc have dots in it.
#things missing: specialization. 

class Sheet:
    #todo:  exports: pdf,  and plain text.
    # character creation: go to the proccess of making a character, and validating it. (7/5/3, 13/9/5, 5will, 7 bg, 15 freebies)


    def __init__(self,name):

        self.name = name
        self.window = root
        root.geometry('800x800')
        vlist= [0,1,2,3,4,5,6,7,8,9]
        wlist= [5,6,7,8,9,10]

        self.d ={}
        c=0
        r=0
        #start header
        #todo: get the object names from a external file, so you can declare them with a loop. solve how to arrange the objects in the gui.
        with open('output.json') as f:
            data = json.load(f)

        c=1
        r=1
        for key,value in data[name].items():
            if type(data[name][key])== dict:#so it gets only the first header itens
                break      
            label= Label(self.window, text = key, background="dark grey")
            label.grid(column= c , row = r)
            c+=1
            self.d[key] = Entry(self.window)
            self.d[key].grid(column = c, row = r)
            self.d[key].delete(0,END) 
            self.d[key].insert(0,data[name][key])
            c+=1
            if c == 7:
                r+=1
                c = 1  
        
        Label(self.window, text= "ATTRIBUTES", background="dark grey").grid(column=4, row=6)
        Label(self.window, text= "Physical", background="dark grey").grid(column=1, row=7)
        Label(self.window, text= "Social", background="dark grey").grid(column=3, row=7)
        Label(self.window, text= "Mental", background="dark grey").grid(column=5, row=7)

        c=1
        r=8
        for key in data[name]["attributes"]["physical"].keys():
            label= Label(self.window, text = key, background="dark grey")
            label.grid(column= c , row = r)
            c+=1
            self.d[key] = ttk.Combobox(self.window, values = vlist, width = 1)
            self.d[key].grid(column = c, row = r)           
            self.d[key].set(data[name]["attributes"]["physical"][key])
            c-=1
            r+=1 

        c=3
        r=8
        for key in data[name]["attributes"]["social"].keys():
            label= Label(self.window, text = key, background="dark grey")
            label.grid(column= c , row = r)
            c+=1
            self.d[key] = ttk.Combobox(self.window, values = vlist, width = 1)
            self.d[key].grid(column = c, row = r)           
            self.d[key].set(data[name]["attributes"]["social"][key])
            c-=1
            r+=1 

        c=5
        r=8
        for key in data[name]["attributes"]["mental"].keys():
            label= Label(self.window, text = key, background="dark grey")
            label.grid(column= c , row = r)
            c+=1
            self.d[key] = ttk.Combobox(self.window, values = vlist, width = 1)
            self.d[key].grid(column = c, row = r)           
            self.d[key].set(data[name]["attributes"]["mental"][key])
            c-=1
            r+=1 
        
        Label(self.window, text= "ABILITIES", background="dark grey").grid(column=4, row=11)
        Label(self.window, text= "Talents", background="dark grey").grid(column=1, row=12)
        Label(self.window, text= "Skills", background="dark grey").grid(column=3, row=12)
        Label(self.window, text= "Knowledges", background="dark grey").grid(column=5, row=12)
        
        priolistAt= [[7,5,3],[13,9,5]]
        l=0 
        c=2
        r=7       
        for key in data[name]["history"]["character creation"].keys():
            self.d[key] = ttk.Combobox(self.window, values = priolistAt[l], width = l+1)
            self.d[key].grid(column = c, row = r)           
            self.d[key].set(data[name]["history"]["character creation"][key])
            c+=2
            if c== 8:
                c=2
                r=12
                l=1


        c=1
        r=13
        for key in data[name]["abilities"]["talents"].keys():
            label= Label(self.window, text = key, background="dark grey")
            label.grid(column= c , row = r)
            c+=1
            self.d[key] = ttk.Combobox(self.window, values = vlist, width = 1)
            self.d[key].grid(column = c, row = r)           
            self.d[key].set(data[name]["abilities"]["talents"][key])
            c-=1
            r+=1            

        c=3
        r=13
        for key in data[name]["abilities"]["skills"].keys():
            label= Label(self.window, text = key, background="dark grey")
            label.grid(column= c , row = r)
            c+=1
            self.d[key] = ttk.Combobox(self.window, values = vlist, width = 1)
            self.d[key].grid(column = c, row = r)           
            self.d[key].set(data[name]["abilities"]["skills"][key])
            c-=1
            r+=1    

        c=5
        r=13
        for key in data[name]["abilities"]["knowledges"].keys():
            label= Label(self.window, text = key, background="dark grey")
            label.grid(column= c , row = r)
            c+=1
            self.d[key] = ttk.Combobox(self.window, values = vlist, width = 1)
            self.d[key].grid(column = c, row = r)           
            self.d[key].set(data[name]["abilities"]["knowledges"][key])
            c-=1
            r+=1   

        c=1
        r=25
        Label(self.window, text= "SPHERES", background="dark grey").grid(column=4, row=24)
        for key,value in data[name]["spheres"].items(): #unique to mage   
            label= Label(self.window, text = key, background="dark grey")
            label.grid(column= c , row = r)
            c+=1
            self.d[key] = ttk.Combobox(self.window, values = vlist, width = 1)
            self.d[key].grid(column = c, row = r)           
            self.d[key].set(value)
            c+=1
            if c == 7:
                r+=1
                c = 1  

        Label(self.window, text= "ADVANTAGES", background="dark grey").grid(column=4, row=28)
        Label(self.window, text= "BACKGROUNDS", background="dark grey").grid(column=1, row=29)
        c=1
        r=30
        for key in data[name]["backgrounds"].keys():# no label, nome dos itens != conteudo
            if c == 1:
                self.d[key] = Entry(self.window)
                self.d[key].grid(column = c, row = r)
                self.d[key].delete(0,END) 
                self.d[key].insert(0,data[name]["backgrounds"][key])                
            elif c ==2:
                self.d[key] = ttk.Combobox(self.window, values = vlist, width = 1)
                self.d[key].grid(column = c, row = r)           
                self.d[key].set(data[name]["backgrounds"][key])
            c+=1
            if c==3:
                c=1
                r+=1
        
        Label(self.window, text= "Arete", background="dark grey").grid(column=3, row=29)  #unique to mage 
        self.d["arete"] = ttk.Combobox(self.window, values = vlist, width = 1)
        self.d["arete"].grid(column = 4, row = 29)           
        self.d["arete"].set(data[name]["arete"])
        Label(self.window, text= "Willpower", background="dark grey").grid(column=3, row=30)  #unique to mage 
        self.d["willpower"] = ttk.Combobox(self.window, values = wlist, width = 1)
        self.d["willpower"].grid(column = 4, row = 30)           
        self.d["willpower"].set(data[name]["willpower"])
        Label(self.window, text= "OTHER TRAITS", background="dark grey").grid(column=3, row=31)  
        c=3
        r=32

        for key in data[name]["other traits"].keys():# no label, nome dos itens != conteudo
            if c == 3:
                self.d[key] = Entry(self.window)
                self.d[key].grid(column = c, row = r)
                self.d[key].delete(0,END) 
                self.d[key].insert(0,data[name]["other traits"][key]) 
            elif c == 4:
                self.d[key] = ttk.Combobox(self.window, values = vlist, width = 1)
                self.d[key].grid(column = c, row = r)           
                self.d[key].set(data[name]["other traits"][key])
            c+=1
            if c==5:
                c=3
                r+=1    

        Label(self.window, text= "MERITS AND FLAWS", background="dark grey").grid(column=5, row=29)
        c=5
        r=30

        for key in data[name]["merits and flaws"].keys():
            if c == 5:
                self.d[key] = Entry(self.window)
                self.d[key].grid(column = c, row = r)
                self.d[key].delete(0,END) 
                self.d[key].insert(0,data[name]["merits and flaws"][key])  
            elif c == 6:
                self.d[key] = ttk.Combobox(self.window, values = vlist, width = 1)
                self.d[key].grid(column = c, row = r)           
                self.d[key].set(data[name]["merits and flaws"][key])
            c+=1
            if c==7:
                c=5
                r+=1
        
        button_erase = Button(self.window, text= "Erase", width = 10, command= self.erase )
        button_erase.grid(column = 1, row =36)
        button_exportExcel = Button(self.window, text= "Export for Excel", width = 10, command= self.exportExcel)
        button_exportExcel.grid(column = 4, row =36)
        button_save = Button(self.window, text= "Save", width= 20, command= self.save)
        button_save.grid( column = 5, row = 36)

    def openSheet(self,name): #deprecated 
        with open('output.json') as f:
            data = json.load(f)

        if name in data.keys():
            for key,value in data[name].items():#this is the loop to get all nested dictionaries, 3 levels in 
                if type(value) == dict:
                    for key1,value1 in value.items():
                        if type(value1) == dict:#aqui começa para os backgrounds
                            for key2 in value1.keys():
                                if type(self.ui.d[key2]) == Entry: 
                                    self.ui.d[key2].delete(0,END)
                                    self.ui.d[key2].insert(0,data[name][key][key1][key2])        
                                else:
                                    self.ui.d[key2].set(data[name][key][key1][key2])                                  
                        else:
                            if type(self.ui.d[key1])== Entry:
                                self.ui.d[key1].delete(0,END)
                                self.ui.d[key1].insert(0,data[name][key][key1])
                            else:
                                self.ui.d[key1].set(data[name][key][key1])   
                else:
                    if type(self.ui.d[key]) == Entry:
                        self.ui.d[key].delete(0,END) 
                        self.ui.d[key].insert(0,data[name][key])
                    else:
                        self.ui.d[key].set(data[name][key])                        
        else:
             messagebox.showerror("Character Not Found", "There is no character named " + name)            

    def sum(self,name,category):#give the sum of points of a certain category.
        sum= 0
        with open('output.json') as f:
            data = json.load(f)

        if name in data.keys():
            if category == "physical" or category == "social" or category=="mental":
                for value in data[name]["attributes"][category].values():
                    sum+= int(value)
            elif category == "talents" or category =="skills" or category == "knowledges":
                for value in data[name]["abilities"][category].values():
                    sum+= int(value)            
            elif category == "backgrounds" or category == "spheres":
                for value in data[name][category].values():
                    sum+= int(value)                                    
            elif category == "merits and flaws" or category == "other traits":
                for value in data[name][category].values():
                    sum+= int(value)   
            else:
                return "invalid category"
            
            with open('output.json', 'w') as f: #dump the changes into the json file
                json.dump(data, f, indent = 2) 

            return sum  
        else:
            return "invalid name"

        with open('output.json', 'w') as f: #dump the changes into the json file
            json.dump(data, f, indent = 2)

    def save(self):#solves error and issues to invoke saveProc()
       
        with open('output.json') as f:
            data = json.load(f)
        if self.d["name"].get() == '0' or self.d["name"].get() == "":
             messagebox.showerror("Invalid Character name.", "Character name cannot be 0 or empty.")
             return
        else: #character doesnt have a 0 as name            
            if self.d["name"].get() in data:
                 res = messagebox.askquestion("Character name already exists.", "There is already one named "+self.d["name"].get() + "\nDo you wish to overwrite it?")
                 if res == 'yes':
                     self.saveProc()
                 elif res == 'no':
                     return      
            else:              

                self.saveProc()

    def saveProc(self):#save into the json file
        with open('output.json') as f:
            data = json.load(f)
        ##### Save procedure #####
        x = ''
        x = self.d["name"].get()
        data[x] = copy.deepcopy(data['0']) # data['0'] is the empty form for the sheets
        #bipty bopty get all of the data from the self.ui and shove into data[nameofthesheet]opty
        #data[x]["name"] = self.d["name"].get()
        for key,value in data[x].items():#this is the loop to get all nested dictionaries, present in sheet, 2 levels in
            if type(value) == dict:
                for k,v in value.items():
                    if type(v) == dict:
                        for p in v.keys():# p é key, x é value
                            data[x][key][k][p] = self.d[p].get()
                    else:
                        data[x][key][k] = self.d[k].get()   
            else:
                data[x][key] = self.d[key].get()  
        with open('output.json', 'w') as f: #dump the changes into the json file
            json.dump(data, f, indent = 2)

    def erase(self):

        if self.d["name"].get() =="0":
            messagebox.showinfo('Response', 'Cannot delete sheet 0')
            return
        res = messagebox.askquestion('askquestion', 'Are you sure you want to erase the sheet?')
        if res == 'yes':
            with open('output.json') as f:
                data = json.load(f)
            del data[self.d["name"].get()]
            with open('output.json', 'w') as f: #dump the changes into the json file
                json.dump(data, f, indent = 2)
            messagebox.showinfo('Response', 'Sheet deleted!')
        elif res == 'no':
            messagebox.showinfo('Response', 'Phew! Nothing was erased.')
        else:
            messagebox.showwarning('error', 'Something went wrong!')           

    def exportExcel(self):
        
        name = self.d["name"].get()
        wb = openpyxl.Workbook()
        tab = wb.active #tab is the name used for the excel sheet, cant be named sheet because sheet is used for character sheet
        tab.title = name    
        with open('output.json') as f:
            data = json.load(f)

        c=1
        r=1
                
        for key,value in data[name].items():
            if type(data[name][key])== dict:
                break      
            cellref = tab.cell(row = r, column = c)
            cellref.value = key + ' :'
            c+=1
            cellref = tab.cell(row = r, column = c)
            cellref.value = value
            c+=1
            if c == 7:
                r+=1
                c = 1  
        
        c=1
        r=6
        for key in data[name]["attributes"]["physical"].keys():
            cellref = tab.cell(row = r, column= c, value = key)
            c+=1
            cellref = tab.cell(row = r, column= c, value = data[name]["attributes"]["physical"][key])
            c-=1
            r+=1 

        c=3
        r=6
        for key in data[name]["attributes"]["social"].keys():
            cellref = tab.cell(row = r, column= c, value = key)
            c+=1
            cellref = tab.cell(row = r, column= c, value = data[name]["attributes"]["social"][key])
            c-=1
            r+=1 

        c=5
        r=6
        for key in data[name]["attributes"]["mental"].keys():
            cellref = tab.cell(row = r, column= c, value = key)
            c+=1
            cellref = tab.cell(row = r, column= c, value = data[name]["attributes"]["mental"][key])
            c-=1
            r+=1 

        c=1
        r=11
        for key in data[name]["abilities"]["talents"].keys():
            cellref = tab.cell(row = r, column= c, value = key)
            c+=1
            cellref = tab.cell(row = r, column= c, value = data[name]["abilities"]["talents"][key])
            c-=1
            r+=1            

        c=3
        r=11
        for key in data[name]["abilities"]["skills"].keys():
            cellref = tab.cell(row = r, column= c, value = key)
            c+=1
            cellref = tab.cell(row = r, column= c, value = data[name]["abilities"]["skills"][key])
            c-=1
            r+=1    

        c=5
        r=11
        for key in data[name]["abilities"]["knowledges"].keys():
            cellref = tab.cell(row = r, column= c, value = key)
            c+=1
            cellref = tab.cell(row = r, column= c, value = data[name]["abilities"]["knowledges"][key])
            c-=1
            r+=1   

        c=1
        r=23

        for key,value in data[name]["spheres"].items(): #unique to mage   
            cellref = tab.cell(row = r, column = c)
            cellref.value = key + ' :'
            c+=1
            cellref = tab.cell(row = r, column = c)
            cellref.value = value
            c+=1
            if c == 7:
                r+=1
                c = 1  

        c=1
        r=29
        for key in data[name]["backgrounds"].keys():
            if data[name]["backgrounds"][key] =="":
                break
            cellref = tab.cell(row = r, column= c, value = data[name]["backgrounds"][key])
            c+=1
            if c==3:
                c=1
                r+=1
        c=3
        r=30
        for key in data[name]["other traits"].keys():
            if data[name]["other traits"][key] =="":
                break
            cellref = tab.cell(row = r, column= c, value = data[name]["other traits"][key])
            c+=1
            if c==5:
                c=3
                r+=1

        c=5
        r=28
        for key in data[name]["merits and flaws"].keys():
            if data[name]["merits and flaws"][key] =="":
                break
            cellref = tab.cell(row = r, column= c, value = data[name]["merits and flaws"][key])
            c+=1
            if c==7:
                c=5
                r+=1

        cellref = tab.cell(row =27 , column =3, value = "arete")  #unique to mage
        cellref = tab.cell(row =27 , column =4, value = data[name]["arete"])   #unique to mage     
        cellref = tab.cell(row =28 , column =3, value = "willpower")  
        cellref = tab.cell(row =28 , column =4, value = data[name]["willpower"]) 
        cellref = tab.cell(row=4 , column=4, value="Attributes")  
        cellref = tab.cell(row=5 , column=1, value="Physical") 
        cellref = tab.cell(row=5 , column=3, value="Social") 
        cellref = tab.cell(row=5 , column=5, value="Mental") 
        cellref = tab.cell(row=9 , column=4, value="Abilities")
        cellref = tab.cell(row=10 , column=1, value="Talents")         
        cellref = tab.cell(row=10 , column=3, value="Skills")         
        cellref = tab.cell(row=10 , column=5, value="Knowledges")
        cellref = tab.cell(row=22 , column=4, value="Spheres")#unique to mage
        cellref = tab.cell(row=26 , column=4, value="Advantages")
        cellref = tab.cell(row=28 , column=1, value="Backgrounds")
        cellref = tab.cell(row=27 , column=5, value="Merits and Flaws")
        cellref = tab.cell(row=29 , column=3, value="Other Traits")

        excel_export_dir= '.\\excel_export\\' 
        if os.path.exists(excel_export_dir) == False:
            os.makedirs(excel_export_dir)

        wb.save(excel_export_dir + name + '_sheet.xlsx')
        messagebox.showinfo('Sheet Exported!','Sheet exported as ' + name + '_sheet.xlsx')

    def unwrap(self, name):#unwrap the json sheet into a unested dict
        sheet = {}
        with open('output.json') as f:
            data = json.load(f)

        for key,value in data[name].items():#this is the loop to get all nested dictionaries, 3 levels in, and unwrap in a unested dict
            if type(value) == dict:
                for key1,value1 in value.items():
                    if type(value1) == dict:
                        for key2 in value1.keys():
                                sheet[key2] = data[name][key][key1][key2]                                  
                    else:
                        sheet[key1] = data[name][key][key1] #bg ot and mf are all recorded here   
            else:
                sheet[key] = data[name][key]      
        return sheet


"""
    def validateNewCharacter(self):
        self.saveProc(self)
        with open('output.json') as f:
            data = json.load(f)       

        self.ui["name"](state='disabled')
        name = ""
        name = data["name"]["name"]
        if self.sum(name, "spheres")<6:
            message.showerror("Missing " +  str(self.sum(name, "spheres") - 6*(-1)) +" points in spheres")
        elif self.sum(name, "spheres")>6:
            message.showerror("Overspent "+ self.sum(name, "spheres")-6 + " dots in spheres")
        if self.sum(name, "backgrounds")<7:
            #message: faltam self.sum(name, "backgrounds") -7*(-1) pts de bg
        elif self.sum(name, "backgrounds")>7:
            #message: self.sum(name, "backgrounds") -7 pts de bg
        priority= {"physical":data["name"]["phyPrio"],"mental":data["name"]["menPrio"],"social":data["name"]["socPrio"]}
        if priority["physical"] == priority["mental"] or priority["physical"]==priority["social"] or priority["mental"]==priority["social"]:
            #message: you cannot have 2 categories with the same priority
        for key in priority.keys():
            if self.sum(name,key)-3 != int(priority[key]):
                if sum < points:
                    #message: faltam points-sum em key
                else:
                    #message: sobram sum-points em key
        priorityAb = {"talents":data["name"]["talPrio"] , "skills":data["name"]["skiPrio"] , "knowledges"data["name"]["knoPrio"]:}
        if priority["physical"] == priority["mental"] or priority["physical"]==priority["social"] or priority["mental"]==priority["social"]:
            #message: you cannot have 2 categories with the same priority
        for key in priorityAb.keys():
            if self.sum(name,key) != int(priorityAb[key]):
                if sum < points:
                    #message: faltam points-sum em key
                else:
                    #message: sobram sum-points em key
        totalFreebiesAtt = (sum(name,"physical") + sum(name,"social") + sum(name,"mental") - 21) *5 #da o total de pontos gastos nos att
        totalFreebiesAbi = (sum(name,"talents")+ sum(name,"skills") + sum(name,"skills") - 27) *2
        totalFreebiesBgs = sum(name,"backgrounds") -7
        totalFreebiesAre = (int(data["name"]["arete"])-1)*4
        totalFreebiesWil = int(data["name"]["will"])-5
        totalFreebiesMfs = sum(name,"merits and flaws")
        totalFreebiesOtt = sum(name, "other traits")*2
        totalFreebiesSph = (sum(name,"spheres")-6)*7

"""
class Window(Frame):
    def __init__(self, master=None):
        Frame.__init__(self, master)
        self.master = master

        menu = Menu(self.master)
        self.master.config(menu=menu)
        with open('output.json') as f:
            data = json.load(f)
        
        fileMenu = Menu(menu, tearoff=False)
        empty = "0"
        fileMenu.add_command(label="New", command = lambda y=empty :Sheet(y))
        fileMenu.add_command(label="Exit", command=self.exitProgram)
        menu.add_cascade(label="File", menu=fileMenu)

        openMenu = Menu(menu, tearoff=False)
        for key in data.keys():
            if key == "0":
                pass
            else:
                openMenu.add_command(label=key, command=lambda y=key :Sheet(y))
        menu.add_cascade(label="Open", menu=openMenu)



    def exitProgram(self):
        exit()
                    
root = Tk()
root.title('Wod Sheet Manager')
root.configure(background= "dark grey")
root.geometry('700x500')
app = Window(root)
root.mainloop()
